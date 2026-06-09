"""Excel (.xlsx) export. Reuses the shared row builders from csv_util so the
spreadsheet columns/rows match the CSV exactly."""
import io

from openpyxl import Workbook
from openpyxl.styles import Font

from csv_util import rows_for

_SHEET_TITLES = {"invoice": "Invoice", "receipt": "Receipt", "statement": "Statement"}


def document_to_xlsx(doc_type: str, data: dict) -> bytes:
    """Build an .xlsx workbook for an extracted document and return the bytes."""
    headers, rows = rows_for(doc_type, data)

    wb = Workbook()
    ws = wb.active
    ws.title = _SHEET_TITLES.get(doc_type, "Data")

    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for row in rows:
        ws.append(["" if v is None else v for v in row])

    # Reasonable auto-ish column widths based on header/content length.
    for col_idx, header in enumerate(headers, start=1):
        max_len = len(str(header))
        for row in rows:
            val = row[col_idx - 1] if col_idx - 1 < len(row) else ""
            max_len = max(max_len, len(str(val if val is not None else "")))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(48, max(10, max_len + 2))

    ws.freeze_panes = "A2"  # keep the header row visible when scrolling

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
